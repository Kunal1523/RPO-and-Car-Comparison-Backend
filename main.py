import requests
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import List, Optional
from DBManager import BrandDbManager, CarDbManager, VariantDbManager,PricingDbManager, FeatureDbManager
from fastapi import UploadFile, File
import openpyxl
import pdb
import logging
import os
import sys
from fastapi.middleware.cors import CORSMiddleware
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv
from DBmanager1 import UserDBHandler
from fastapi import FastAPI,Request, Depends, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from sqlalchemy import inspect, Table
import time

import models, schemas
from database import engine, get_db
from dependencies import get_current_user
import os
from dotenv import load_dotenv
from fastapi import FastAPI,APIRouter, HTTPException,Depends,BackgroundTasks
import requests
from fastapi.responses import RedirectResponse, JSONResponse
from pydantic import BaseModel
import os
from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode
import time
import httpx
load_dotenv()


SERPAPI_KEY = os.getenv("SERPAPI_KEY")
userdbhandler = UserDBHandler()


CLIENT_ID = os.getenv("MICROSOFT_CLIENT_ID")
CLIENT_SECRET = os.getenv("MICROSOFT_CLIENT_SECRET")
REDIRECT_URI = os.getenv("MICROSOFT_REDIRECT_URI")
TENANT_ID = os.getenv("TENANT_ID")
SCOPE = os.getenv("MICROSOFT_SCOPE")
# AUTH_URL = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/authorize"
# TOKEN_URL = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
AUTH_URL = f"https://login.microsoftonline.com/common/oauth2/v2.0/authorize"
TOKEN_URL = f"https://login.microsoftonline.com/common/oauth2/v2.0/token"
USERINFO_URL = "https://graph.microsoft.com/v1.0/me"
MICROSOFT_GRAPH_URL = "https://graph.microsoft.com/v1.0/me/sendMail"
# ---------------------------
# DEV schema guard
# ---------------------------
def ensure_schema(engine):
    """
    Dev-only schema guard:
    - Creates tables if missing
    - If tables exist but are missing required columns -> drops only those tables and recreates
    """
    inspector = inspect(engine)

    required = {
        "drafts": {"id", "name", "owner_email", "updated_at", "data"},
        "final_plans": {"owner_email", "published_at", "published_by", "data"},
        "regulations": {"id", "owner_email", "name"},
        "models": {"id", "owner_email", "name"},
        "tenant_users": {"email", "name", "tenant_id", "is_active"},
        "user_settings": {"owner_email", "years_window", "start_month", "preferences"},
        "audit_log": {"id", "owner_email", "action", "created_at", "details"},
    }

    existing_tables = set(inspector.get_table_names())

    # If any required table is missing, create everything
    if any(t not in existing_tables for t in required.keys()):
        print("[DB] One or more tables missing -> creating all tables.")
        models.Base.metadata.create_all(bind=engine)
        return

    # Identify broken tables (missing columns)
    bad_tables = []
    for table_name, required_cols in required.items():
        cols = {c["name"] for c in inspector.get_columns(table_name)}
        missing = required_cols - cols
        if missing:
            bad_tables.append((table_name, missing))

    if not bad_tables:
        print("[DB] Schema OK.")
        return

    # Drop only broken tables
    print("[DB] Schema mismatch detected. Dropping broken tables (DEV MODE):")
    for table_name, missing in bad_tables:
        print(f" - {table_name} missing columns: {sorted(missing)}")

    metadata = models.Base.metadata
    with engine.begin() as conn:
        for table_name, _ in bad_tables:
            print(f"[DB] Dropping table: {table_name}")
            Table(table_name, metadata, autoload_with=engine).drop(conn, checkfirst=True)

    # Recreate tables (dropped ones will be recreated)
    print("[DB] Recreating tables...")
    models.Base.metadata.create_all(bind=engine)
    print("[DB] Done.")


# Run schema check at import/startup time
ensure_schema(engine)

def get_loggers():
    LOG_DIR = os.path.abspath("logs")
    os.makedirs(LOG_DIR, exist_ok=True)

    def _setup(name, filename):
        logger = logging.getLogger(name)
        logger.setLevel(logging.INFO)
        logger.propagate = False

        if not logger.handlers:
            handler = logging.FileHandler(
                os.path.join(LOG_DIR, filename),
                mode="a",   # ✅ APPEND, NOT WRITE
                encoding="utf-8"
            )
            formatter = logging.Formatter(
                "%(asctime)s | %(levelname)s | %(message)s"
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)

        return logger

    return (
        _setup("valid_rows", "valid_rows.log"),
        _setup("duplicate_rows", "duplicate_rows.log"),
        _setup("skipped_rows", "skipped_rows.log"),
    )

app = FastAPI(title="Car Compare API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*","http://localhost:3000", "http://localhost:5173","https://kaylyn-unoppugned-vertie.ngrok-free.dev"],  # In production, specify your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

brand_db = BrandDbManager()
car_db = CarDbManager()
variant_db = VariantDbManager()
pricing_db = PricingDbManager()
feature_db = FeatureDbManager()


class FeatureUpdateRequest(BaseModel):
    variant_id: str
    feature_id: str
    value: str
    version: int = 1

class BrandCreateRequest(BaseModel):
    name: str


class CarCreateRequest(BaseModel):
    brand_name: str
    car_name: str

class VariantBulkCreateRequest(BaseModel):
    brand_name: str
    car_name: str
    variants: list[str]


# class PricingItem(BaseModel):
#     variant_id: str
#     ex_showroom_price: float
#     currency: str = "INR"

class PricingItem(BaseModel):
    variant_id: str
    type: str = "Standard"       # <<--- add this
    ex_showroom_price: float
    currency: str = "INR"

class BulkPricingRequest(BaseModel):
    version: int = 1
    pricing: List[PricingItem]



class FeatureCreate(BaseModel):
    name: str
    category: str


class BulkFeatureCreateRequest(BaseModel):
    features: List[FeatureCreate]


class PricingQueryRequest(BaseModel):
    variant_id: str
    version: int


class VariantCompareRequest(BaseModel):
    variant_ids: List[str]
    version: int = 1

class PriceInsertRequest(BaseModel):
    variant_id: int
    price: float
    type: str




def write_audit(db: Session, owner_email: str, action: str, details: dict):
    db.add(
        models.AuditLog(
            owner_email=owner_email,
            action=action,
            created_at=int(time.time() * 1000),
            details=details or {},
        )
    )
    db.commit()

SPECIAL_VALUES = ["Deadline NM", "Deadline AM"]

def extract_models_from_plan(plan_data: dict) -> List[str]:
    """
    plan_data is the JSON stored in FinalPlan.data
    expected shape: {"regulationCells": { "Reg 1|FY 25-26|4": ["A","B", ...], ... } }
    """
    cells = (plan_data or {}).get("regulationCells", {}) or {}
    found = set()
    for values in cells.values():
        if not isinstance(values, list):
            continue
        for v in values:
            if not v:
                continue
            v = str(v).strip()
            if not v or v in SPECIAL_VALUES:
                continue
            found.add(v)
    return sorted(found)

def extract_regulations_from_plan(plan_data: dict) -> List[str]:
    """
    Extract regulation rowIds from keys: "RegName|FY 25-26|4"
    """
    cells = (plan_data or {}).get("regulationCells", {}) or {}
    regs = set()
    for k in cells.keys():
        if not isinstance(k, str):
            continue
        parts = k.split("|")
        if len(parts) >= 1 and parts[0].strip():
            regs.add(parts[0].strip())
    return sorted(regs)

def upsert_user_regulations(db: Session, owner_email: str, regs_list: List[str]):
    db.query(models.Regulation).filter(models.Regulation.owner_email == owner_email).delete()
    for name in regs_list:
        db.add(models.Regulation(owner_email=owner_email, name=name))
    db.commit()

def upsert_user_models(db: Session, owner_email: str, models_list: List[str]):
    # simplest: replace-all for that user
    db.query(models.ModelItem).filter(models.ModelItem.owner_email == owner_email).delete()
    for name in models_list:
        db.add(models.ModelItem(owner_email=owner_email, name=name))
    db.commit()


def compute_compliance(db: Session, owner_email: str, plan_data: dict) -> dict:
    """
    Returns: { "Reg 1": ["B","C"], "Reg 2": ["A"], ... } only missing regs
    Uses DB regulations + DB models (frozen on publish)
    """
    regs = db.query(models.Regulation).filter(models.Regulation.owner_email == owner_email).all()
    regs_names = [r.name for r in regs]

    model_rows = db.query(models.ModelItem).filter(models.ModelItem.owner_email == owner_email).all()
    all_models = [m.name for m in model_rows]

    if not regs_names or not all_models:
        return {}

    cells = (plan_data or {}).get("regulationCells", {}) or {}

    missing_by_reg = {}
    for reg in regs_names:
        planned = set()
        prefix = f"{reg}|"
        for key, values in cells.items():
            if not isinstance(key, str) or not key.startswith(prefix):
                continue
            if not isinstance(values, list):
                continue
            for v in values:
                v = str(v).strip()
                if v and v not in SPECIAL_VALUES:
                    planned.add(v)

        missing = [m for m in all_models if m not in planned]
        if missing:
            missing_by_reg[reg] = missing

    return missing_by_reg

# ---------------------------
# DRAFTS
# ---------------------------
class DraftRename(BaseModel):
    name: str


@app.get("/health")
def health():
    return {"ok": True}

@app.post("/brands")
def create_brand(payload: BrandCreateRequest):
    try:
        result = brand_db.insert_brand(payload.name)
        return {"success": True, "data": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/cars")
def create_car(payload: CarCreateRequest):
    try:
        brand_id = brand_db.get_brand_id_by_name(payload.brand_name)

        if not brand_id:
            raise HTTPException(
                status_code=404,
                detail="Brand not found. Please insert brand first."
            )

        result = car_db.insert_car(brand_id, payload.car_name)
        return {"success": True, "data": result}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/catalog")
def get_catalog():
    try:
        rows = variant_db.get_catalog_data()

        result = {}
        for r in rows:
            brand = result.setdefault(r["brand_id"], {
                "brand_id": r["brand_id"],
                "brand_name": r["brand_name"],
                "cars": {}
            })

            car = brand["cars"].setdefault(r["car_id"], {
                "car_id": r["car_id"],
                "car_name": r["car_name"],
                "variants": []
            })

            car["variants"].append({
                "variant_id": r["variant_id"],
                "variant_name": r["variant_name"],
                "version": r["version"]
            })

        # convert dicts → lists
        response = {
            "brands": [
                {
                    **b,
                    "cars": list(b["cars"].values())
                }
                for b in result.values()
            ]
        }

        return response

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.post("/api/pricing")
def get_pricing(payload: PricingQueryRequest):
    try:
        price = pricing_db.get_price(
            variant_id=payload.variant_id,
            version=payload.version,
        )

        if not price:
            raise HTTPException(
                status_code=404,
                detail="Pricing not found for selected inputs"
            )

        return {
            "variant_id": payload.variant_id,
            "version": payload.version,
            "type": price["type"],
            "ex_showroom_price": price["ex_showroom_price"],
            "currency": price["currency"]
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/variants/bulk")
def create_variants(payload: VariantBulkCreateRequest):
    try:
        car_id = variant_db.get_car_id(payload.brand_name, payload.car_name)

        if not car_id:
            raise HTTPException(status_code=404, detail="Car not found")

        result = variant_db.bulk_insert_variants(car_id, payload.variants)

        return {
            "success": True,
            "inserted_variants": result
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/variants")
def get_variants(brand_name: str, car_name: str):
    """
    Example:
    /variants?brand_name=Maruti Suzuki&car_name=Grand Vitara
    """
    try:
        variants = variant_db.get_variants_by_brand_and_car(
            brand_name=brand_name,
            car_name=car_name
        )

        if not variants:
            return {
                "success": False,
                "message": "No variants found"
            }

        return {
            "success": True,
            "brand": brand_name,
            "car": car_name,
            "variants": variants
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/pricing/bulk")
def bulk_insert_pricing(payload: BulkPricingRequest):
    try:
        return pricing_db.bulk_insert_pricing(
            pricing_list=[p.dict() for p in payload.pricing],
            version=payload.version
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pricing")
def get_pricing(brand_name: str, car_name: str):
    try:
        pricing_data = pricing_db.get_pricing_by_brand_and_car(
            brand_name=brand_name,
            car_name=car_name
        )

        if not pricing_data:
            return {
                "success": False,
                "message": "No pricing data found"
            }

        prices = [p["ex_showroom_price"] for p in pricing_data]

        return {
            "success": True,
            "brand": brand_name,
            "car": car_name,
            "summary": {
                "min_price": min(prices),
                "max_price": max(prices),
                "total_variants": len(pricing_data),
                "currency": pricing_data[0].get("currency", "INR")
            },
            "pricing": pricing_data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/pricing")
def get_pricing(brand_name: str, car_name: str):
    try:
        pricing_data = pricing_db.get_pricing_by_brand_and_car_v1(
            brand_name=brand_name,
            car_name=car_name
        )

        if not pricing_data:
            return {
                "success": False,
                "message": "No pricing data found"
            }

        return {
            "success": True,
            "brand": brand_name,
            "car": car_name,
            "total_configurations": len(pricing_data),
            "pricing": pricing_data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    

    

@app.post("/features/bulk")
def bulk_insert_features(payload: BulkFeatureCreateRequest):
    pdb.set_trace()
    try:
        result = feature_db.bulk_insert_features(
            features=[f.dict() for f in payload.features]
        )

        return {
            "success": True,
            "data": result
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/new-features/master/from-excel")
def insert_new_features_from_excel(file: UploadFile = File(...)):
    """
    Upload final reviewed Excel.
    Inserts ONLY NEW features into features_master.
    IGNORE → skipped
    MAPPED → skipped
    """
    try:
        wb = openpyxl.load_workbook(file.file)
        sheet = wb.active

        headers = [cell.value for cell in sheet[1]]
        header_index = {h: i for i, h in enumerate(headers)}

        required_cols = {"original_name", "final_action", "Category"}
        if not required_cols.issubset(header_index.keys()):
            raise HTTPException(
                status_code=400,
                detail=f"Excel must contain columns: {required_cols}"
            )

        CATEGORY_NORMALIZATION = {
            "Infotainment and Connectivity": "Infotainment",
            "Infotainment & Connectivity": "Infotainment",
        }

        new_features = []
        skipped = 0

        for row in sheet.iter_rows(min_row=2):
            action = row[header_index["final_action"]].value
            if action != "NEW":
                skipped += 1
                continue

            raw_name = row[header_index["original_name"]].value
            raw_category = row[header_index["Category"]].value

            if not raw_name or not raw_category:
                continue

            category = CATEGORY_NORMALIZATION.get(
                raw_category.strip(),
                raw_category.strip()
            )

            new_features.append({
                "name": raw_name.strip(),
                "category": category
            })

        if not new_features:
            return {
                "success": True,
                "message": "No NEW features found",
                "inserted": 0
            }

        # 🔐 DB layer should already handle UNIQUE constraint
        result = feature_db.bulk_insert_features(new_features)

        return {
            "success": True,
            "inserted": len(result),
            "skipped_rows": skipped,
            "data": result
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/features/master/category-wise")
def get_feature_master_category_wise():
    try:
        return feature_db.get_feature_master_category_wise()
    except Exception as e:
        return {"error": str(e)}
    

@app.post("/features/master/normalize")
def normalize_feature_master():
    try:
        result = feature_db.normalize_feature_master()
        return {
            "success": True,
            "message": "Feature master normalized successfully",
            "data": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/normalize-variants-and-pricing")
def normalize_variants_and_pricing(brand_name: str, car_name: str):
    try:
        conn = variant_db.conn

        with conn.cursor() as cursor:

            # 1️⃣ Get car_id
            cursor.execute("""
                SELECT c.id
                FROM cars c
                JOIN brands b ON b.id = c.brand_id
                WHERE b.name = %s AND c.name = %s
            """, (brand_name, car_name))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(404, "Car not found")
            car_id = row[0]

            # 2️⃣ Get all variants of this car
            cursor.execute("""
                SELECT id, name
                FROM variants
                WHERE car_id = %s
            """, (car_id,))
            variants = cursor.fetchall()

            master_map = {}  # clean_name -> master_variant_id
            fixed_variants = 0
            moved_prices = 0

            for variant_id, name in variants:
                lower = name.lower()

                # Detect paint type
                if "dual" in lower:
                    paint_type = "dual_tone"
                elif "metallic" in lower:
                    paint_type = "metallic"
                else:
                    paint_type = "standard"

                # Normalize variant name
                clean_name = (
                    name.replace("(Dual Tone)", "")
                        .replace("(Metallic)", "")
                        .strip()
                )

                # 3️⃣ Get / create master variant
                if clean_name not in master_map:
                    cursor.execute("""
                        SELECT id FROM variants
                        WHERE car_id = %s AND name = %s
                    """, (car_id, clean_name))
                    existing = cursor.fetchone()

                    if existing:
                        master_id = existing[0]
                    else:
                        cursor.execute("""
                            INSERT INTO variants (car_id, name)
                            VALUES (%s, %s)
                            RETURNING id
                        """, (car_id, clean_name))
                        master_id = cursor.fetchone()[0]

                    master_map[clean_name] = master_id
                else:
                    master_id = master_map[clean_name]

                # 4️⃣ Move pricing
                cursor.execute("""
                    UPDATE pricing
                    SET variant_id = %s,
                        paint_type = %s
                    WHERE variant_id = %s
                """, (master_id, paint_type, variant_id))

                if variant_id != master_id:
                    # 5️⃣ Delete duplicate variant
                    cursor.execute("""
                        DELETE FROM variants WHERE id = %s
                    """, (variant_id,))
                    fixed_variants += 1

                moved_prices += cursor.rowcount

        conn.commit()

        return {
            "success": True,
            "variants_removed": fixed_variants,
            "pricing_rows_updated": moved_prices,
            "final_variants_expected": "≈20",
            "message": "Variants normalized & pricing separated by paint type"
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))

# import re

# CATEGORY_ALIAS_MAP = {
#     "suzuki connect": "connected car technology"
# }


# def normalize_category(text: str) -> str:
#     if not text:
#         return ""
#     key = text.strip().lower()
#     return CATEGORY_ALIAS_MAP.get(key, key)


# def normalize_feature(text: str) -> str:
#     """
#     Normalize feature name for matching
#     Example:
#       'Displacement' -> 'displacement'
#       'Max Power (KW)' -> 'max power'
#     """
#     text = text.lower().strip()
#     text = re.sub(r"\(.*?\)", "", text)   # remove units/parentheses
#     text = re.sub(r"\s+", " ", text)
#     return text


# @app.post("/variant-features/upload-excel")
# def upload_variant_features_excel(
#     brand_name: str,
#     car_name: str,
#     file: UploadFile = File(...)
# ):
#     conn = feature_db.conn
#     skipped_rows = []

#     try:
#         wb = openpyxl.load_workbook(file.file)
#         sheet = wb.active

#         with conn.cursor() as cursor:

#             # 1️⃣ Get car_id
#             cursor.execute("""
#                 SELECT c.id
#                 FROM cars c
#                 JOIN brands b ON b.id = c.brand_id
#                 WHERE b.name = %s AND c.name = %s
#             """, (brand_name, car_name))
#             row = cursor.fetchone()
#             if not row:
#                 raise HTTPException(404, "Car not found")
#             car_id = row[0]

#             # 2️⃣ Variant map
#             cursor.execute("""
#                 SELECT id, name
#                 FROM variants
#                 WHERE car_id = %s
#             """, (car_id,))
#             variant_map = {
#                 name.strip().lower(): vid
#                 for vid, name in cursor.fetchall()
#             }

#             # 3️⃣ Feature master map (NORMALIZED)
#             cursor.execute("""
#                 SELECT id, name, category
#                 FROM features_master
#             """)
#             feature_map = {
#                 (
#                     normalize_feature(name),
#                     category.strip().lower()
#                 ): fid
#                 for fid, name, category in cursor.fetchall()
#             }

#             inserted = 0
#             skipped = 0

#             # 4️⃣ Read Excel rows
#             # 4️⃣ Read Excel rows
#             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
#                 category, feature_name, variant_name, value = row

#                 reason = None

#                 if not category or not feature_name or not variant_name or value is None:
#                     reason = "Missing data"
#                 else:
#                     # Variant lookup
#                     v_key = variant_name.strip().lower()
#                     if v_key not in variant_map:
#                         reason = f"Variant not found: '{variant_name}'"

#                     # ✅ Feature lookup with CATEGORY ALIAS
#                     f_key = (
#                         normalize_feature(feature_name),
#                         normalize_category(category)
#                     )

#                     if not reason and f_key not in feature_map:
#                         reason = f"Feature not found: '{feature_name}' in category '{category}'"

#                 if reason:
#                     skipped += 1
#                     skipped_rows.append({
#                         "row_number": idx,
#                         "category": category,
#                         "feature_name": feature_name,
#                         "variant_name": variant_name,
#                         "value": value,
#                         "reason": reason
#                     })
#                     continue

#                 variant_id = variant_map[v_key]
#                 feature_id = feature_map[f_key]

#                 cursor.execute("""
#                     INSERT INTO variant_features
#                         (variant_id, feature_id, value)
#                     VALUES (%s, %s, %s)
#                     ON CONFLICT (variant_id, feature_id, version)
#                     DO UPDATE SET
#                         value = EXCLUDED.value,
#                         is_latest = true
#                 """, (
#                     variant_id,
#                     feature_id,
#                     str(value).strip()
#                 ))

#                 inserted += 1


#         conn.commit()

#         return {
#             "success": True,
#             "inserted": inserted,
#             "skipped": skipped,
#             "skipped_rows": skipped_rows,  # <-- detailed info
#             "message": "Excel processed safely (feature mapped, no new insert)"
#         }

#     except Exception as e:
#         conn.rollback()
#         raise HTTPException(500, str(e))


from psycopg2.extras import execute_values
import openpyxl
import re
import logging
import os
from fastapi import UploadFile, File, HTTPException

# -------------------------------
# LOG DIRECTORY SETUP
# -------------------------------


# -------------------------------
# NORMALIZATION
# -------------------------------
CATEGORY_ALIAS_MAP = {
    "suzuki connect": "connected car technology"
}

def normalize_category(text: str) -> str:
    if not text:
        return ""
    return CATEGORY_ALIAS_MAP.get(text.strip().lower(), text.strip().lower())

def normalize_feature(text: str) -> str:
    text = text.lower().strip()
    # text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"\s+", " ", text)
    return text

# -------------------------------
# API
# -------------------------------
@app.post("/variant-features/upload-excel-fast")
def upload_variant_features_excel_fast(
    brand_name: str,
    car_name: str,
    file: UploadFile = File(...)
):
    valid_logger, duplicate_logger, skipped_logger = get_loggers()
    valid_logger.info("=== API CALLED ===")
    conn = feature_db.conn
    skipped_rows = []

    try:
        wb = openpyxl.load_workbook(file.file, read_only=True)
        sheet = wb.active

        # -------------------------------
        # 1️⃣ LOAD MASTER DATA
        # -------------------------------
        with conn.cursor() as cursor:

            cursor.execute("""
                SELECT c.id
                FROM cars c
                JOIN brands b ON b.id = c.brand_id
                WHERE b.name = %s AND c.name = %s
            """, (brand_name, car_name))
            car = cursor.fetchone()
            if not car:
                raise HTTPException(404, "Car not found")
            car_id = car[0]

            cursor.execute("""
                SELECT id, name
                FROM variants
                WHERE car_id = %s
            """, (car_id,))
            variant_map = {name.strip().lower(): vid for vid, name in cursor.fetchall()}

            cursor.execute("""
                SELECT id, name, category
                FROM features_master
            """)
            feature_map = {
                (normalize_feature(name), category.strip().lower()): fid
                for fid, name, category in cursor.fetchall()
            }

        # -------------------------------
        # 2️⃣ PROCESS EXCEL ROWS
        # -------------------------------
        bulk_rows = []
        seen_keys = set()

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            category, feature_name, variant_name, value = row
            reason = None

            if not category or not feature_name or not variant_name or value is None:
                reason = "Missing mandatory field"
            else:
                v_key = variant_name.strip().lower()
                if v_key not in variant_map:
                    reason = f"Variant not found: {variant_name}"

                f_key = (normalize_feature(feature_name), normalize_category(category))
                if not reason and f_key not in feature_map:
                    reason = f"Feature not found: {feature_name} ({category})"

            if reason:
                skipped_logger.info(
                    f"Row {idx} | Category={category} | Feature={feature_name} | "
                    f"Variant={variant_name} | Value={value} | Reason={reason}"
                )
                skipped_rows.append({
                    "row_number": idx,
                    "reason": reason
                })
                continue

            variant_id = variant_map[v_key]
            feature_id = feature_map[f_key]
            key = (variant_id, feature_id)

            if key in seen_keys:
                duplicate_logger.info(
                    f"Row {idx} | DUPLICATE | variant_id={variant_id} | "
                    f"feature_id={feature_id} | Feature={feature_name}"
                )
                continue

            seen_keys.add(key)

            valid_logger.info(
                f"Row {idx} | variant_id={variant_id} | feature_id={feature_id} | "
                f"Value={value} | Feature={feature_name}"
            )

            bulk_rows.append((
                variant_id,
                feature_id,
                str(value).strip(),
                feature_name.strip()
            ))

        # -------------------------------
        # 3️⃣ INSERT
        # -------------------------------
        if not bulk_rows:
            return {"success": False, "message": "No valid rows to insert"}

        with conn.cursor() as cursor:
            execute_values(
                cursor,
                """
                INSERT INTO variant_features
                    (variant_id, feature_id, value, original_name)
                VALUES %s
                ON CONFLICT (variant_id, feature_id, version)
                DO UPDATE SET
                    value = EXCLUDED.value,
                    original_name = EXCLUDED.original_name,
                    is_latest = true
                """,
                bulk_rows,
                page_size=1000
            )

        conn.commit()
        
        for logger in (valid_logger, duplicate_logger, skipped_logger):
            for h in logger.handlers:
                h.flush()

        return {
            "success": True,
            "inserted": len(bulk_rows),
            "skipped": len(skipped_rows),
            "logs": {
                "valid": "logs/valid_rows.log",
                "duplicates": "logs/duplicate_rows.log",
                "skipped": "logs/skipped_rows.log"
            }
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))




# ============== MAIN COMPARISON API (Compatible with your frontend) ==============

# @app.post("/api/compare/variants")
# def compare_variants(payload: VariantCompareRequest):
#     """
#     Compare multiple variants - Returns data in format compatible with ComparisonTable.tsx
    
#     Response format:
#     {
#         "columns": ["Feature", "Variant 1", "Variant 2", ...],
#         "data": [
#             {"feature": "Price Value", "Variant 1": "10 Lakh", "Variant 2": "12 Lakh"},
#             {"feature": "Safety - Airbags", "Variant 1": "6", "Variant 2": "6"},
#             ...
#         ]
#     }
#     """
#     try:
#         if len(payload.variant_ids) < 2:
#             raise HTTPException(
#                 status_code=400,
#                 detail="At least 2 variants required for comparison"
#             )
        
#         if len(payload.variant_ids) > 5:
#             raise HTTPException(
#                 status_code=400,
#                 detail="Maximum 5 variants can be compared at once"
#             )

#         # Step 1: Get variant details and build columns
#         columns = ["Feature"]
#         variant_details = []
        
#         for variant_id in payload.variant_ids:
#             variant_info = variant_db.get_variant_details(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             if not variant_info:
#                 raise HTTPException(
#                     status_code=404,
#                     detail=f"Variant {variant_id} not found"
#                 )
            
#             # Column name format: "Brand Car - Variant"
#             column_name = f"{variant_info['brand_name']} {variant_info['car_name']} - {variant_info['variant_name']}"
#             columns.append(column_name)
#             variant_details.append({
#                 "variant_id": variant_id,
#                 "column_name": column_name,
#                 "info": variant_info
#             })
        
#         # Step 2: Build data rows
#         data_rows = []
        
#         # Add Price & Launch Info rows
#         for variant_detail in variant_details:
#             variant_id = variant_detail["variant_id"]
            
#             # Get pricing
#             pricing = pricing_db.get_price(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             # Store pricing for this variant
#             variant_detail["pricing"] = pricing
        
#         # Price Value row
#         price_row = {"feature": "Price Value"}
#         for variant_detail in variant_details:
#             col = variant_detail["column_name"]
#             pricing = variant_detail["pricing"]
            
#             if pricing and pricing.get("ex_showroom_price"):
#                 # Format price in lakhs
#                 price = float(pricing["ex_showroom_price"])
#                 price_lakhs = price / 100000
#                 price_row[col] = f"₹ {price_lakhs:.2f} Lakh"
#             else:
#                 price_row[col] = ""
        
#         data_rows.append(price_row)
        
#         # Variant Launched row (if launch_year exists)
#         launch_row = {"feature": "Variant Launched"}
#         has_launch_data = False
#         for variant_detail in variant_details:
#             col = variant_detail["column_name"]
#             launch_year = variant_detail["info"].get("launch_year")
            
#             if launch_year:
#                 launch_row[col] = str(launch_year)
#                 has_launch_data = True
#             else:
#                 launch_row[col] = ""
        
#         if has_launch_data:
#             data_rows.append(launch_row)
        
#         # Step 3: Get all features and organize by category
#         all_features = {}  # feature_id -> {name, category, values_by_variant}
        
#         for variant_detail in variant_details:
#             variant_id = variant_detail["variant_id"]
#             col = variant_detail["column_name"]
            
#             features = feature_db.get_variant_features(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             for feature in features:
#                 feature_id = feature["feature_id"]
                
#                 if feature_id not in all_features:
#                     all_features[feature_id] = {
#                         "name": feature["feature_name"],
#                         "category": feature["category"],
#                         "values": {}
#                     }
                
#                 # Store value for this variant
#                 value = feature["value"] if feature["value"] else ""
#                 all_features[feature_id]["values"][col] = value
        
#         # Step 4: Create feature rows in format "Category - Feature Name"
#         for feature_id, feature_data in all_features.items():
#             row = {
#                 "feature": f"{feature_data['category']} - {feature_data['name']}"
#             }
            
#             # Add value for each variant
#             for variant_detail in variant_details:
#                 col = variant_detail["column_name"]
#                 row[col] = feature_data["values"].get(col, "")
            
#             data_rows.append(row)
        
#         # Step 5: Return response
#         return {
#             "columns": columns,
#             "data": data_rows
#         }
    
#     except HTTPException:
#         raise
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         raise HTTPException(status_code=500, detail=str(e))

# @app.post("/api/compare/variants")
# def compare_variants(payload: VariantCompareRequest):
#     """
#     Compare multiple variants - Returns data in format compatible with ComparisonTable.tsx
    
#     Response format includes pricing breakdown for hover tooltips
#     """
#     try:
#         if len(payload.variant_ids) < 2:
#             raise HTTPException(
#                 status_code=400,
#                 detail="At least 2 variants required for comparison"
#             )
        
#         if len(payload.variant_ids) > 5:
#             raise HTTPException(
#                 status_code=400,
#                 detail="Maximum 5 variants can be compared at once"
#             )

#         # Step 1: Get variant details and build columns
#         columns = ["Feature"]
#         variant_details = []
#         variant_pricing_data = {}  # ✅ Store complete pricing breakdown for tooltips
        
#         for variant_id in payload.variant_ids:
#             variant_info = variant_db.get_variant_details(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             if not variant_info:
#                 raise HTTPException(
#                     status_code=404,
#                     detail=f"Variant {variant_id} not found"
#                 )
            
#             # Column name format: "Brand Car - Variant"
#             column_name = f"{variant_info['brand_name']} {variant_info['car_name']} - {variant_info['variant_name']}"
#             columns.append(column_name)
            
#             # ✅ Get ALL prices for this variant (all types)
#             all_prices = pricing_db.get_all_prices(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             # ✅ Build pricing data structure for tooltip
#             prices_list = []
#             total_price = 0
#             price_count = 0
            
#             for price_info in all_prices:
#                 if price_info.get("ex_showroom_price"):
#                     price_value = float(price_info["ex_showroom_price"])
#                     price_lakhs = price_value / 100000
                    
#                     prices_list.append({
#                         "type": price_info.get("type", "standard"),
#                         "currency": price_info.get("currency", "INR"),
#                         "ex_showroom_price": price_value,
#                         "price_display": f"₹ {price_lakhs:.2f} Lakh"
#                     })
                    
#                     total_price += price_value
#                     price_count += 1
            
#             # Calculate average price
#             avg_price = None
#             if price_count > 0:
#                 avg_value = total_price / price_count
#                 avg_lakhs = avg_value / 100000
#                 avg_price = {
#                     "value": avg_value,
#                     "display": f"₹ {avg_lakhs:.2f} Lakh"
#                 }
            
#             # Store complete pricing data for this variant
#             variant_pricing_data[column_name] = {
#                 "variant_id": variant_id,
#                 "variant_name": variant_info['variant_name'],
#                 "prices": prices_list,
#                 "avg_price": avg_price
#             }
            
#             variant_details.append({
#                 "variant_id": variant_id,
#                 "column_name": column_name,
#                 "info": variant_info,
#                 "pricing_data": variant_pricing_data[column_name]
#             })
        
#         # Step 2: Build data rows
#         data_rows = []
        
#         # Price Value row (using average price)
#         price_row = {"feature": "Price Value"}
#         for variant_detail in variant_details:
#             col = variant_detail["column_name"]
#             pricing_data = variant_detail["pricing_data"]
            
#             if pricing_data and pricing_data.get("avg_price"):
#                 price_row[col] = pricing_data["avg_price"]["display"]
#             else:
#                 price_row[col] = ""
        
#         data_rows.append(price_row)
        
#         # Variant Launched row (if launch_year exists)
#         launch_row = {"feature": "Variant Launched"}
#         has_launch_data = False
#         for variant_detail in variant_details:
#             col = variant_detail["column_name"]
#             launch_year = variant_detail["info"].get("launch_year")
            
#             if launch_year:
#                 launch_row[col] = str(launch_year)
#                 has_launch_data = True
#             else:
#                 launch_row[col] = ""
        
#         if has_launch_data:
#             data_rows.append(launch_row)
        
#         # Step 3: Get all features and organize by category
#         all_features = {}  # feature_id -> {name, category, values_by_variant}
        
#         for variant_detail in variant_details:
#             variant_id = variant_detail["variant_id"]
#             col = variant_detail["column_name"]
            
#             features = feature_db.get_variant_features(
#                 variant_id=variant_id,
#                 version=payload.version
#             )
            
#             for feature in features:
#                 feature_id = feature["feature_id"]
                
#                 if feature_id not in all_features:
#                     all_features[feature_id] = {
#                         "name": feature["feature_name"],
#                         "category": feature["category"],
#                         "values": {}
#                     }
                
#                 # Store value for this variant
#                 value = feature["value"] if feature["value"] else ""
#                 all_features[feature_id]["values"][col] = value
        
#         # Step 4: Create feature rows in format "Category - Feature Name"
#         for feature_id, feature_data in all_features.items():
#             row = {
#                 "feature": f"{feature_data['category']} - {feature_data['name']}"
#             }
            
#             # Add value for each variant
#             for variant_detail in variant_details:
#                 col = variant_detail["column_name"]
#                 row[col] = feature_data["values"].get(col, "")
            
#             data_rows.append(row)
        
#         # ✅ Step 5: Return response with pricing data for tooltips
#         return {
#             "columns": columns,
#             "data": data_rows,
#             "variant_pricing": variant_pricing_data  # ✅ Complete pricing breakdown
#         }
    
#     except HTTPException:
#         raise
#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/compare/variants")
def compare_variants(payload: VariantCompareRequest):
    """
    Compare multiple variants
    Pricing is embedded inside 'Price Value' row under data section
    """
    try:
        if len(payload.variant_ids) < 2:
            raise HTTPException(400, "At least 2 variants required")

        if len(payload.variant_ids) > 5:
            raise HTTPException(400, "Maximum 5 variants allowed")

        columns = ["Feature"]
        variant_details = []

        # -------------------------
        # Step 1: Variant details + pricing
        # -------------------------
        for variant_id in payload.variant_ids:
            variant_info = variant_db.get_variant_details(
                variant_id=variant_id,
                version=payload.version
            )

            if not variant_info:
                raise HTTPException(404, f"Variant {variant_id} not found")

            column_name = f"{variant_info['brand_name']} {variant_info['car_name']} - {variant_info['variant_name']}"
            columns.append(column_name)

            all_prices = pricing_db.get_all_prices(
                variant_id=variant_id,
                version=payload.version
            )

            prices_list = []
            total = 0
            count = 0

            for p in all_prices:
                if p.get("ex_showroom_price"):
                    price_val = float(p["ex_showroom_price"])
                    # price_lakh = price_val / 100000

                    prices_list.append({
                        "currency": p.get("currency", "INR"),
                        "ex_showroom_price": price_val,
                        # "display": f"₹ {price_lakh:.2f} Lakh",
                        "fuel_type":p.get("fuel_type"),
                        "engine_type":p.get("engine_type"),
                        "transmission_type":p.get("transmission_type"),
                        "paint_type":p.get("paint_type"),
                        "edition":p.get("edition")
                    })

                    total += price_val
                    count += 1

            # avg_price = None
            # if count > 0:
            #     avg_val = total / count
            #     avg_price = {
            #         "value": avg_val,
            #         "display": f"₹ {avg_val / 100000:.2f} Lakh"
            #     }

            variant_details.append({
                "variant_id": variant_id,
                "column": column_name,
                "info": variant_info,
                "pricing": {
                    "prices": prices_list
                    # "avg_price": avg_price
                }
            })

        # -------------------------
        # Step 2: Build data rows
        # -------------------------
        data_rows = []

        # ✅ PRICE VALUE ROW (pricing embedded here)
        price_row = {"feature": "Price Value"}

        for v in variant_details:
            if v["pricing"]["prices"]:
                price_row[v["column"]] = {
                    "pricing": v["pricing"]   # 👈 FULL pricing inside data
                }
            else:
                price_row[v["column"]] = None

        data_rows.append(price_row)

        # Variant Launch Year
        launch_row = {"feature": "Variant Launched"}
        has_launch = False

        for v in variant_details:
            year = v["info"].get("launch_year")
            if year:
                launch_row[v["column"]] = str(year)
                has_launch = True
            else:
                launch_row[v["column"]] = ""

        if has_launch:
            data_rows.append(launch_row)

        # -------------------------
        # Step 3: Features
        # -------------------------
        all_features = {}

        for v in variant_details:
            features = feature_db.get_variant_features(
                variant_id=v["variant_id"],
                version=payload.version
            )

            for f in features:
                fid = f["feature_id"]
                if fid not in all_features:
                    all_features[fid] = {
                        "name": f["feature_name"],
                        "category": f["category"],
                        "values": {}
                    }

                all_features[fid]["values"][v["column"]] = f.get("value", "")

        for f in all_features.values():
            row = {
                "feature": f"{f['category']} - {f['name']}"
            }
            for v in variant_details:
                row[v["column"]] = f["values"].get(v["column"], "")
            data_rows.append(row)

        # -------------------------
        # Final Response
        # -------------------------
        return {
            "columns": columns,
            "data": data_rows
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))

# update feature value API endpoint
@app.put("/api/variant/feature/update")
def update_variant_feature(payload: FeatureUpdateRequest):

    try:

        updated = feature_db.update_variant_feature(
            variant_id=payload.variant_id,
            feature_id=payload.feature_id,
            value=payload.value,
            version=payload.version
        )

        if not updated:
            raise HTTPException(
                404,
                "Variant or Feature not found"
            )

        return {
            "status": "success",
            "message": "Feature updated successfully"
        }

    except HTTPException:
        raise

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, str(e))

    
# ============== HELPER API: Get Variant IDs ==============

@app.get("/api/variants/search")
def search_variants(
    brand_name: Optional[str] = None,
    car_name: Optional[str] = None,
    limit: int = 20
):
    """
    Search for variants to get their IDs for comparison
    Returns list with variant_id, names, and pricing
    """
    try:
        variants = variant_db.get_comparable_variants(
            brand_name=brand_name,
            car_name=car_name,
            limit=limit
        )
        
        return {
            "success": True,
            "variants": variants
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============== HELPER API: Get Brands & Cars for Dropdowns ==============

@app.get("/api/brands-cars")
def get_brands_and_cars():
    """
    Get all brands with their cars for building selection UI
    """
    try:
        query = """
            SELECT 
                b.id as brand_id,
                b.name as brand_name,
                c.id as car_id,
                c.name as car_name
            FROM brands b
            LEFT JOIN cars c ON b.id = c.brand_id
            ORDER BY b.name, c.name
        """
        
        with variant_db.conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query)
            results = cur.fetchall()
        
        # Group by brand
        brands_map = {}
        for row in results:
            brand_id = row["brand_id"]
            
            if brand_id not in brands_map:
                brands_map[brand_id] = {
                    "brand_id": brand_id,
                    "brand_name": row["brand_name"],
                    "cars": []
                }
            
            if row["car_id"]:
                brands_map[brand_id]["cars"].append({
                    "car_id": row["car_id"],
                    "car_name": row["car_name"]
                })
        
        return {
            "success": True,
            "brands": list(brands_map.values())
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============== HELPER API: Get Variants by Car ==============

# @app.get("/api/cars/{car_id}/variants")
# def get_variants_by_car(car_id: str):
#     """
#     Get all variants for a specific car with pricing
#     """
#     try:
#         query = """
#             SELECT 
#                 v.id as variant_id,
#                 v.name as variant_name,
#                 v.version,
#                 c.name as car_name,
#                 b.name as brand_name,
#                 p.ex_showroom_price,
#                 p.currency,
#                 p.type
#             FROM variants v
#             JOIN cars c ON v.car_id = c.id
#             JOIN brands b ON c.brand_id = b.id
#             LEFT JOIN pricing p ON v.id = p.variant_id AND v.version = p.version
#             WHERE c.id = %s AND v.is_latest = true
#             ORDER BY p.ex_showroom_price NULLS LAST
#         """
        
#         with variant_db.conn.cursor(cursor_factory=RealDictCursor) as cur:
#             cur.execute(query, (car_id,))
#             results = cur.fetchall()
        
#         if not results:
#             raise HTTPException(status_code=404, detail="Car not found")
        
#         variants = []
#         for row in results:
#             variant = dict(row)
            
#             # Format price for display
#             if variant["ex_showroom_price"]:
#                 price_lakhs = float(variant["ex_showroom_price"]) / 100000
#                 variant["price_display"] = f"₹ {price_lakhs:.2f} Lakh"
#             else:
#                 variant["price_display"] = "Price not available"
            
#             variants.append(variant)
        
#         return {
#             "success": True,
#             "brand_name": results[0]["brand_name"],
#             "car_name": results[0]["car_name"],
#             "variants": variants
#         }
    
#     except HTTPException:
#         raise
#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))
    


@app.get("/api/cars/{car_id}/variants")
def get_variants_by_car(car_id: str):
    """
    Get all variants for a specific car with all pricing types + avg price
    """
    try:
        query = """
            SELECT 
                v.id as variant_id,
                v.name as variant_name,
                v.version,
                c.name as car_name,
                b.name as brand_name,
                p.ex_showroom_price,
                p.currency,
                p.type
            FROM variants v
            JOIN cars c ON v.car_id = c.id
            JOIN brands b ON c.brand_id = b.id
            LEFT JOIN pricing p 
                ON v.id = p.variant_id 
                AND v.version = p.version
            WHERE c.id = %s 
              AND v.is_latest = true
            ORDER BY v.name, p.ex_showroom_price NULLS LAST
        """

        with variant_db.conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(query, (car_id,))
            rows = cur.fetchall()

        if not rows:
            raise HTTPException(status_code=404, detail="Car not found")

        brand_name = rows[0]["brand_name"]
        car_name = rows[0]["car_name"]

        variant_map = {}

        for row in rows:
            vid = row["variant_id"]

            if vid not in variant_map:
                variant_map[vid] = {
                    "variant_id": vid,
                    "variant_name": row["variant_name"],
                    "version": row["version"],
                    "prices": [],
                    "avg_price": None
                }

            # Add price if exists
            if row["ex_showroom_price"] is not None:
                price_value = float(row["ex_showroom_price"])
                price_lakhs = price_value / 100000

                variant_map[vid]["prices"].append({
                    "type": row["type"],
                    "currency": row["currency"],
                    "ex_showroom_price": price_value,
                    "price_display": f"₹ {price_lakhs:.2f} Lakh"
                })

        # compute avg price for each variant
        for variant in variant_map.values():
            prices = [p["ex_showroom_price"] for p in variant["prices"]]

            if prices:
                avg_price = sum(prices) / len(prices)
                avg_price_lakhs = avg_price / 100000

                variant["avg_price"] = {
                    "value": avg_price,
                    "display": f"₹ {avg_price_lakhs:.2f} Lakh"
                }
            else:
                variant["avg_price"] = None

        return {
            "success": True,
            "brand_name": brand_name,
            "car_name": car_name,
            "variants": list(variant_map.values())
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/news")
def get_car_news(car: str):
    if not SERPAPI_KEY:
        raise HTTPException(status_code=500, detail="SerpAPI key missing")

    # SerpAPI Google News endpoint
    url = "https://serpapi.com/search.json"

    params = {
        "engine": "google_news",
        "q": f"{car} car latest features news",
        "gl": "in",                        # India region
        "hl": "en",                        # English
        "num": 10,                         # Fetch extra to filter manually
        "api_key": SERPAPI_KEY,
    }

    response = requests.get(url, params=params)
    data = response.json()

    if "news_results" not in data:
        raise HTTPException(status_code=500, detail="Failed to fetch news")

    articles = data["news_results"]

    # Filter out motorsport/racing junk if present
    clean_articles = []
    for item in articles:
        title = item.get("title", "").lower()
        if any(x in title for x in ["wrc", "rally", "motorsport", "race"]):
            continue
        clean_articles.append(item)

        if len(clean_articles) == 5:
            break

    formatted = [
        {
            "title": a.get("title"),
            "description": a.get("snippet"),
            "url": a.get("link"),
            "source": a.get("source"),
            "published": a.get("date")
        }
        for a in clean_articles
    ]

    return {
        "car": car,
        "total": len(formatted),
        "top5_news": formatted
    }


 ## Admin Level APIs
class UpdateVariantFeature(BaseModel):
    variant_id: str
    feature_id: str
    value: Optional[str]


class CreateFeature(BaseModel):
    name: str
    category: str
    is_active: Optional[bool] = True


@app.get("/admin/brands")
def get_brands():
    try:
        return brand_db.get_all_brands()
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/cars")
def get_cars(brand_id: str):
    try:
        return car_db.get_cars_by_brand_id(brand_id)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/variants")
def get_variants(car_id: str):
    try:
        return variant_db.get_variants_by_car_id(car_id)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/features")
def get_features(category: Optional[str] = None):
    try:
        return feature_db.get_features(category)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/variant-feature")
def get_variant_feature(variant_id: str, feature_id: str):
    try:
        row = feature_db.get_variant_feature_latest(variant_id, feature_id)

        return row if row else {"value": None}
    except Exception as e:
        raise HTTPException(500, str(e))

@app.put("/admin/variant-feature/update")
def update_variant_feature(payload: UpdateVariantFeature):
    try:
        feature_db.update_variant_feature_value(
            variant_id=payload.variant_id,
            feature_id=payload.feature_id,
            value=payload.value
        )

        return {"success": True, "message": "Feature updated"}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/admin/variant/features")
def get_all_features_for_variant(variant_id: str):
    try:
        return feature_db.get_all_features_for_variant(variant_id)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.get("/admin/categories")
def get_categories():
    try:
        return feature_db.get_categories()
    except Exception as e:
        raise HTTPException(500, str(e))



@app.post("/admin/features")
def add_feature(payload: FeatureCreate):
    try:
        return feature_db.create_feature(
            name=payload.name,
            category=payload.category
        )
    except Exception as e:
        raise HTTPException(400, str(e))



@app.get("/pricing/by-ids")
def get_pricing_by_ids(brand_id: int, car_id: int):
    try:
        query = """
        SELECT
            v.id AS variant_id,
            v.name AS variant_name,
            p.ex_showroom_price,
            p.currency,
            p.version AS pricing_version,
            p.type,
            p.effective_date
        FROM pricing p
        JOIN variants v ON p.variant_id = v.id
        JOIN cars c ON v.car_id = c.id
        WHERE
            c.brand_id = %s
            AND c.id = %s
            AND v.is_latest = true
            AND p.is_latest = true
        ORDER BY p.ex_showroom_price DESC;
        """
        
        with pricing_db.conn.cursor() as cursor:
            cursor.execute(query, (brand_id, car_id))
            rows = cursor.fetchall()

        if not rows:
            return {
                "success": False,
                "message": "No pricing data found"
            }

        pricing_data = [
            {
                "variant_id": row[0],
                "variant_name": row[1],
                "ex_showroom_price": float(row[2]) if row[2] else None,
                "currency": row[3],
                "pricing_version": row[4],
                "type": row[5],
                "effective_date": row[6].isoformat() if row[6] else None
            }
            for row in rows
        ]

        prices = [p["ex_showroom_price"] for p in pricing_data if p["ex_showroom_price"]]

        return {
            "success": True,
            "brand_id": brand_id,
            "car_id": car_id,
            "summary": {
                "min_price": min(prices) if prices else 0,
                "max_price": max(prices) if prices else 0,
                "total_variants": len(pricing_data),
                "currency": pricing_data[0].get("currency", "INR") if pricing_data else "INR"
            },
            "pricing": pricing_data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/pricing")
def get_pricing(brand_name: str, car_name: str):
    try:
        pricing_data = pricing_db.get_pricing_by_brand_and_car(
            brand_name=brand_name,
            car_name=car_name
        )

        if not pricing_data:
            return {
                "success": False,
                "message": "No pricing data found"
            }

        prices = [p["ex_showroom_price"] for p in pricing_data]

        return {
            "success": True,
            "brand": brand_name,
            "car": car_name,
            "summary": {
                "min_price": min(prices),
                "max_price": max(prices),
                "total_variants": len(pricing_data),
                "currency": pricing_data[0].get("currency", "INR")
            },
            "pricing": pricing_data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/v1/admin/pricing")
def get_pricing(brand_name: str, car_name: str):
    try:
        pricing_data = pricing_db.get_pricing_by_brand_and_car_v1(
            brand_name=brand_name,
            car_name=car_name
        )

        if not pricing_data:
            return {
                "success": False,
                "message": "No pricing data found"
            }

        return {
            "success": True,
            "brand": brand_name,
            "car": car_name,
            "total_configurations": len(pricing_data),
            "pricing": pricing_data
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.patch("/admin/pricing/update")
def update_price(variant_id: int, new_price: float):
    """Updates only the price field for the existing latest record."""
    try:
        success = pricing_db.update_existing_price(variant_id, new_price)
        if not success:
            raise HTTPException(status_code=404, detail="Variant not found or no latest price to update")
        
        return {"success": True, "message": f"Price updated for variant {variant_id}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/admin/pricing/insert")
def insert_price(data: PriceInsertRequest):
    """Inserts a brand new price record for a variant."""
    try:
        pricing_db.insert_new_price(
            variant_id=data.variant_id, 
            price=data.price, 
            p_type=data.type
        )
        return {"success": True, "message": "New price record inserted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))






# Project Planning APIs


@app.get("/drafts", response_model=List[schemas.DraftBase])
def get_drafts(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    drafts = (
        db.query(models.Draft)
        .filter(models.Draft.owner_email == current_user_email)
        .order_by(models.Draft.updated_at.desc())
        .all()
    )
    return drafts


@app.get("/drafts/{draft_id}", response_model=schemas.DraftBase)
def get_draft(
    draft_id: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    d = db.query(models.Draft).filter(models.Draft.id == draft_id).first()
    if not d:
        raise HTTPException(404, "Draft not found")
    if d.owner_email != current_user_email:
        raise HTTPException(403, "Not allowed")
    return d


@app.post("/drafts", response_model=schemas.DraftBase)
def create_draft(
    draft: schemas.DraftCreate,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    # Upsert logic by ID, enforce ownership
    existing = db.query(models.Draft).filter(models.Draft.id == draft.id).first()

    if existing:
        if existing.owner_email != current_user_email:
            raise HTTPException(status_code=403, detail="Not authorized to modify this draft")

        existing.name = draft.name
        existing.updated_at = draft.updatedAt
        existing.data = draft.data.model_dump()
        db.commit()
        db.refresh(existing)
        return existing

    new_draft = models.Draft(
        id=draft.id,
        name=draft.name,
        owner_email=current_user_email,
        updated_at=draft.updatedAt,
        data=draft.data.model_dump(),
    )
    db.add(new_draft)
    db.commit()
    db.refresh(new_draft)
    return new_draft


@app.patch("/drafts/{draft_id}/rename", response_model=schemas.DraftBase)
def rename_draft(
    draft_id: str,
    payload: DraftRename,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    d = db.query(models.Draft).filter(models.Draft.id == draft_id).first()
    if not d:
        raise HTTPException(404, "Draft not found")
    if d.owner_email != current_user_email:
        raise HTTPException(403, "Not allowed")

    d.name = payload.name
    db.commit()
    db.refresh(d)
    return d


@app.delete("/drafts/{draft_id}")
def delete_draft(
    draft_id: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    d = db.query(models.Draft).filter(models.Draft.id == draft_id).first()
    if not d:
        raise HTTPException(status_code=404, detail="Draft not found")
    if d.owner_email != current_user_email:
        raise HTTPException(status_code=403, detail="Not authorized to delete this draft")

    db.delete(d)
    db.commit()
    return {"ok": True}


# ---------------------------
# FINAL PLAN (per-user, never deleted)
# ---------------------------
class ShareRequest(BaseModel):
    emails: List[str]
    subject: str
    body: str


@app.get("/final-plan", response_model=schemas.FinalPlanOut)
def get_final_plan(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    plan = (
        db.query(models.FinalPlan)
        .filter(models.FinalPlan.owner_email == current_user_email)
        .first()
    )
    if not plan:
        raise HTTPException(status_code=404, detail="No final plan published")

    missing = compute_compliance(db, current_user_email, plan.data)

    # return combined response (single call)
    return {
        "owner_email": plan.owner_email,
        "published_at": plan.published_at,
        "published_by": plan.published_by,
        "data": plan.data,
        "missingByReg": missing
    }



@app.post("/final-plan", response_model=schemas.FinalPlanOut)
def publish_plan(
    plan: schemas.FinalPlanCreate,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    existing = (
        db.query(models.FinalPlan)
        .filter(models.FinalPlan.owner_email == current_user_email)
        .first()
    )

    payload_data = plan.data.model_dump()

    # 1) Upsert final plan
    if existing:
        existing.published_at = plan.publishedAt
        existing.published_by = current_user_email
        existing.data = payload_data
        db.commit()
        db.refresh(existing)
    else:
        existing = models.FinalPlan(
            owner_email=current_user_email,
            published_at=plan.publishedAt,
            published_by=current_user_email,
            data=payload_data,
        )
        db.add(existing)
        db.commit()
        db.refresh(existing)

    # 2) ✅ Save regulations list ONLY on publish (replace all)
    regs = extract_regulations_from_plan(payload_data)
    upsert_user_regulations(db, current_user_email, regs)

    # 3) ✅ Freeze models from plan (replace all)
    frozen_models = extract_models_from_plan(payload_data)
    upsert_user_models(db, current_user_email, frozen_models)

    # 4) ✅ Compute final compliance (from DB regs+models)
    missing = compute_compliance(db, current_user_email, payload_data)

    # 5) Audit
    write_audit(
        db,
        owner_email=current_user_email,
        action="PUBLISH_FINAL",
        details={
            "publishedAt": plan.publishedAt,
            "modelsCount": len(frozen_models),
            "regsCount": len(regs),
        },
    )

    # 6) ✅ Return single object
    return {
        "owner_email": existing.owner_email,
        "published_at": existing.published_at,
        "published_by": existing.published_by,
        "data": existing.data,
        "missingByReg": missing,
    }


@app.post("/final-plan/share")
async def share_final_plan(
    share_data: ShareRequest,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    
    plan = (
        db.query(models.FinalPlan)
        .filter(models.FinalPlan.owner_email == current_user_email)
        .first()
    )
    if not plan:
        raise HTTPException(status_code=404, detail="No final plan to share")

    from .auth_utils import refresh_access_token
    from .email_service import send_email_via_graph, create_plan_excel
    excel_bytes = await create_plan_excel(plan.data)

    tokens = userdbhandler.get_microsoft_tokens(current_user_email)

    if not tokens:
        raise HTTPException(status_code=401, detail="Microsoft account not linked")

    access_token = tokens["access_token"]
    refresh_token = tokens["refresh_token"]

    try:
        await send_email_via_graph(
            access_token=access_token,
            recipients=share_data.emails,
            subject=share_data.subject,
            body=share_data.body,
            attachment_bytes=excel_bytes,
        )
    except Exception:
        # 🔄 refresh & retry once
        refreshed = await refresh_access_token(refresh_token)

        userdbhandler.update_microsoft_tokens(
            current_user_email,
            refreshed["access_token"],
            refreshed["refresh_token"],
            refreshed["expires_at"],
        )

        await send_email_via_graph(
            access_token=refreshed["access_token"],
            recipients=share_data.emails,
            subject=share_data.subject,
            body=share_data.body,
            attachment_bytes=excel_bytes,
        )


# @app.post("/final-plan/share")
# async def share_final_plan(
#     share_data: ShareRequest,
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
#     authorization: Optional[str] = Header(None),
# ):
#     """
#     Share the *current user's* final plan.
#     Uses Microsoft Graph API with user's delegated token.
#     """
#     if not authorization or not authorization.startswith("Bearer "):
#         raise HTTPException(status_code=401, detail="Missing Bearer token for Graph API")

#     token = authorization.split(" ")[1]

#     plan = (
#         db.query(models.FinalPlan)
#         .filter(models.FinalPlan.owner_email == current_user_email)
#         .first()
#     )
#     if not plan:
#         raise HTTPException(status_code=404, detail="No final plan to share")

#     # Generate Excel + send
#     from .email_service import create_plan_excel, send_email_via_graph

#     excel_bytes = await create_plan_excel(plan.data)

#     try:
#         await send_email_via_graph(
#             token=token,
#             recipients=share_data.emails,
#             subject=share_data.subject,
#             body=share_data.body,
#             attachment_bytes=excel_bytes,
#         )
#     except Exception as e:
#         print(f"Graph Send Failed: {e}")
#         raise HTTPException(status_code=500, detail=str(e))

#     # Audit
#     write_audit(
#         db,
#         owner_email=current_user_email,
#         action="SHARE_FINAL",
#         details={"recipients": share_data.emails, "subject": share_data.subject},
#     )

#     return {"ok": True, "sent_to": share_data.emails}


# ---------------------------
# REGULATIONS (per-user)
# ---------------------------
# @app.get("/regulations", response_model=List[str])
# def get_regulations(
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
# ):
#     regs = (
#         db.query(models.Regulation)
#         .filter(models.Regulation.owner_email == current_user_email)
#         .all()
#     )
#     return [r.name for r in regs]
@app.get("/regulations", response_model=List[str])
def get_regulations(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    regs = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.is_archived == False  # <--- Only active ones
        )
        .all()
    )
    return [r.name for r in regs]

# get Models 
@app.get("/models", response_model=List[str])
def get_models(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    models_list = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.is_archived == False,
        )
        .all()
    )

    return [m.name for m in models_list]

# @app.post("/regulations")
# def add_regulation(
#     reg: schemas.RegulationItem,
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
# ):
#     existing = (
#         db.query(models.Regulation)
#         .filter(
#             models.Regulation.owner_email == current_user_email,
#             models.Regulation.name == reg.name,
#         )
#         .first()
#     )
#     if existing:
#         return {"msg": "Already exists"}

#     db.add(models.Regulation(owner_email=current_user_email, name=reg.name))
#     db.commit()
#     return {"ok": True, "name": reg.name}

@app.post("/regulations")
def add_regulation(
    reg: schemas.RegulationItem,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    existing = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.name == reg.name,
            models.Regulation.is_archived == False,   # Only check active ones
        )
        .first()
    )

    if existing:
        return {"msg": "Already exists"}

    new_reg = models.Regulation(
        owner_email=current_user_email,
        name=reg.name,
        is_archived=False,   # ✅ Explicit default
    )

    db.add(new_reg)
    db.commit()
    db.refresh(new_reg)

    return {
        "ok": True,
        "id": new_reg.id,
        "name": new_reg.name,
        "is_archived": new_reg.is_archived,
    }

# ARCHIVE MODELS AND REGULATIONS 
# @app.post("/regulations/archive")
# def archive_regulation(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     # Try to find existing item
#     item = db.query(models.Regulation).filter(
#         models.Regulation.owner_email == current_user_email, 
#         models.Regulation.name == name
#     ).first()
    
#     if not item:
#         # If it doesn't exist in master list, create it as ARCHIVED
#         item = models.Regulation(
#             owner_email=current_user_email,
#             name=name,
#             is_archived=True
#         )
#         db.add(item)
#     else:
#         # If it exists, just flip the bit
#         item.is_archived = True
        
#     db.commit()
#     return {"success": True, "message": f"{name} archived"}




@app.post("/regulations/archive")
def archive_regulation(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    # Find existing active regulation
    item = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.name == name,
            models.Regulation.is_archived == False,   # only active
        )
        .first()
    )

    # If not found → error (do NOT create)
    if not item:
        raise HTTPException(
            status_code=404,
            detail="Regulation not found or already archived"
        )

    # Archive it
    item.is_archived = True
    db.commit()
    db.refresh(item)

    return {
        "success": True,
        "name": item.name,
        "is_archived": item.is_archived,
        "message": "Archived successfully",
    }

# @app.post("/models/archive")
# def archive_model(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.ModelItem).filter(
#         models.ModelItem.owner_email == current_user_email, 
#         models.ModelItem.name == name
#     ).first()
    
#     if not item:
#         item = models.ModelItem(
#             owner_email=current_user_email,
#             name=name,
#             is_archived=True
#         )
#         db.add(item)
#     else:
#         item.is_archived = True
        
#     db.commit()
#     return {"success": True, "message": f"{name} archived"}


@app.post("/models/archive")
def archive_model(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    # Find existing active model
    item = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.name == name,
            models.ModelItem.is_archived == False,   # only active
        )
        .first()
    )

    # If not found → do NOT create
    if not item:
        raise HTTPException(
            status_code=404,
            detail="Model not found or already archived"
        )

    # Archive it
    item.is_archived = True
    db.commit()
    db.refresh(item)

    return {
        "success": True,
        "name": item.name,
        "is_archived": item.is_archived,
        "message": "Archived successfully",
    }
# DELETE MODELS AND REGULATION FROM ARCHIVE 
# Permanent Delete for Models
# @app.delete("/models/permanent")
# def delete_model_permanent(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.ModelItem).filter(
#         models.ModelItem.owner_email == current_user_email, 
#         models.ModelItem.name == name
#     ).first()
    
#     if not item:
#         raise HTTPException(status_code=404, detail="Model not found")
        
#     db.delete(item)
#     db.commit()
#     return {"success": True, "message": f"'{name}' deleted permanently from database"}


@app.delete("/models/permanent")
def delete_model_permanent(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    item = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.name == name,
            models.ModelItem.is_archived == True,   # ✅ only archived
        )
        .first()
    )

    if not item:
        raise HTTPException(
            status_code=404,
            detail="Archived model not found"
        )

    db.delete(item)
    db.commit()

    return {
        "success": True,
        "message": f"'{name}' permanently deleted"
    }
# Permanent Delete for Regulations
# @app.delete("/regulations/permanent")
# def delete_regulation_permanent(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.Regulation).filter(
#         models.Regulation.owner_email == current_user_email, 
#         models.Regulation.name == name
#     ).first()
    
#     if not item:
#         raise HTTPException(status_code=404, detail="Regulation not found")
        
#     db.delete(item)
#     db.commit()
#     return {"success": True, "message": f"'{name}' deleted permanently from database"}
@app.delete("/regulations/permanent")
def delete_regulation_permanent(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    item = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.name == name,
            models.Regulation.is_archived == True,   # ✅ only archived
        )
        .first()
    )

    if not item:
        raise HTTPException(
            status_code=404,
            detail="Archived regulation not found"
        )

    db.delete(item)
    db.commit()

    return {
        "success": True,
        "message": f"'{name}' permanently deleted"
    }
# @app.post("/models/archive")
# def archive_model(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.ModelItem).filter(
#         models.ModelItem.owner_email == current_user_email, 
#         models.ModelItem.name == name
#     ).first()
#     if not item:
#         raise HTTPException(status_code=404, detail="Model not found")
#     item.is_archived = True
#     db.commit()
#     return {"success": True, "message": f"{name} moved to archive"}

# @app.post("/regulations/archive")
# def archive_regulation(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.Regulation).filter(
#         models.Regulation.owner_email == current_user_email, 
#         models.Regulation.name == name
#     ).first()
#     if not item:
#         raise HTTPException(status_code=404, detail="Regulation not found")
#     item.is_archived = True
#     db.commit()
#     return {"success": True, "message": f"{name} moved to archive"}


# VIEW ARCHIVED ITEMS 
# --- VIEW ARCHIVED ITEMS ---
# @app.get("/models/archived", response_model=List[str])
# def get_archived_models(db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     rows = db.query(models.ModelItem).filter(
#         models.ModelItem.owner_email == current_user_email,
#         models.ModelItem.is_archived == True
#     ).all()
#     return [r.name for r in rows]

@app.get("/models/archived", response_model=List[str])
def get_archived_models(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    models_list = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.is_archived == True,
        )
        .all()
    )

    return [m.name for m in models_list]


# @app.get("/regulations/archived", response_model=List[str])
# def get_archived_regulations(db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     rows = db.query(models.Regulation).filter(
#         models.Regulation.owner_email == current_user_email,
#         models.Regulation.is_archived == True
#     ).all()
#     return [r.name for r in rows]
@app.get("/regulations/archived", response_model=List[str])
def get_archived_regulations(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    regs = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.is_archived == True,
        )
        .all()
    )

    return [r.name for r in regs]

# --- RESTORE FROM ARCHIVE ---
# @app.post("/models/restore")
# def restore_model(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.ModelItem).filter(models.ModelItem.owner_email == current_user_email, models.ModelItem.name == name).first()
#     item.is_archived = False
#     db.commit()
#     return {"success": True, "message": f"{name} restored to master list"}
@app.post("/models/restore")
def restore_model(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    item = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.name == name,
            models.ModelItem.is_archived == True,   # ✅ only archived
        )
        .first()
    )

    if not item:
        raise HTTPException(
            status_code=404,
            detail="Archived model not found"
        )

    item.is_archived = False
    db.commit()
    db.refresh(item)

    return {
        "success": True,
        "name": item.name,
        "is_archived": item.is_archived,
        "message": "Model restored successfully",
    }

# @app.post("/regulations/restore")
# def restore_regulation(name: str, db: Session = Depends(get_db), current_user_email: str = Depends(get_current_user)):
#     item = db.query(models.Regulation).filter(models.Regulation.owner_email == current_user_email, models.Regulation.name == name).first()
#     item.is_archived = False
#     db.commit()
#     return {"success": True, "message": f"{name} restored to master list"}

@app.post("/regulations/restore")
def restore_regulation(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    item = (
        db.query(models.Regulation)
        .filter(
            models.Regulation.owner_email == current_user_email,
            models.Regulation.name == name,
            models.Regulation.is_archived == True,   # ✅ only archived
        )
        .first()
    )

    if not item:
        raise HTTPException(
            status_code=404,
            detail="Archived regulation not found"
        )

    item.is_archived = False
    db.commit()
    db.refresh(item)

    return {
        "success": True,
        "name": item.name,
        "is_archived": item.is_archived,
        "message": "Regulation restored successfully",
    }
# @app.delete("/regulations/{name}")
# def delete_regulation(
#     name: str,
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
# ):
#     reg = (
#         db.query(models.Regulation)
#         .filter(
#             models.Regulation.owner_email == current_user_email,
#             models.Regulation.name == name,
#         )
#         .first()
#     )
#     if reg:
#         db.delete(reg)
#         db.commit()
#     return {"ok": True}


# ---------------------------
# MODELS (per-user list used for compliance)
# ---------------------------
# @app.get("/models", response_model=List[str])
# def get_models(
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
# ):
#     rows = (
#         db.query(models.ModelItem)
#         .filter(models.ModelItem.owner_email == current_user_email)
#         .all()
#     )
#     return [r.name for r in rows]

@app.get("/models", response_model=List[str])
def get_models(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    rows = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.is_archived == False  # <--- Only active ones
        )
        .all()
    )
    return [r.name for r in rows] 


# @app.post("/models")
# def add_model(
#     payload: schemas.ModelItemCreate,
#     db: Session = Depends(get_db),
#     current_user_email: str = Depends(get_current_user),
# ):
#     exists = (
#         db.query(models.ModelItem)
#         .filter(
#             models.ModelItem.owner_email == current_user_email,
#             models.ModelItem.name == payload.name,
#         )
#         .first()
#     )
#     if exists:
#         return {"msg": "Already exists"}

#     db.add(models.ModelItem(owner_email=current_user_email, name=payload.name))
#     db.commit()
#     return {"ok": True, "name": payload.name}
@app.post("/models")
def add_model(
    payload: schemas.ModelItemCreate,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):

    # Check only active models
    exists = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.name == payload.name,
            models.ModelItem.is_archived == False,   # only active
        )
        .first()
    )

    if exists:
        return {"msg": "Already exists"}

    new_model = models.ModelItem(
        owner_email=current_user_email,
        name=payload.name,
        is_archived=False,   # ✅ Explicit default
    )

    db.add(new_model)
    db.commit()
    db.refresh(new_model)

    return {
        "ok": True,
        "id": new_model.id,
        "name": new_model.name,
        "is_archived": new_model.is_archived,
    }


@app.delete("/models/{name}")
def delete_model(
    name: str,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    row = (
        db.query(models.ModelItem)
        .filter(
            models.ModelItem.owner_email == current_user_email,
            models.ModelItem.name == name,
        )
        .first()
    )
    if row:
        db.delete(row)
        db.commit()
    return {"ok": True}


# ---------------------------
# USER SETTINGS (per-user)
# ---------------------------
@app.get("/me/settings", response_model=schemas.UserSettingsOut)
def get_settings(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    s = (
        db.query(models.UserSettings)
        .filter(models.UserSettings.owner_email == current_user_email)
        .first()
    )
    if not s:
        s = models.UserSettings(
            owner_email=current_user_email,
            years_window=3,
            start_month=4,
            preferences={},
        )
        db.add(s)
        db.commit()
        db.refresh(s)
    return s


@app.put("/me/settings", response_model=schemas.UserSettingsOut)
def update_settings(
    payload: schemas.UserSettingsUpdate,
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    s = (
        db.query(models.UserSettings)
        .filter(models.UserSettings.owner_email == current_user_email)
        .first()
    )
    if not s:
        s = models.UserSettings(
            owner_email=current_user_email,
            years_window=3,
            start_month=4,
            preferences={},
        )
        db.add(s)
        db.commit()
        db.refresh(s)

    if payload.yearsWindow is not None:
        s.years_window = payload.yearsWindow
    if payload.startMonth is not None:
        s.start_month = payload.startMonth
    if payload.preferences is not None:
        s.preferences = payload.preferences

    db.commit()
    db.refresh(s)
    return s


# ---------------------------
# AUDIT LOG (per-user)
# ---------------------------
@app.get("/audit-log", response_model=List[schemas.AuditLogOut])
def get_audit_log(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
):
    rows = (
        db.query(models.AuditLog)
        .filter(models.AuditLog.owner_email == current_user_email)
        .order_by(models.AuditLog.created_at.desc())
        .limit(100)
        .all()
    )
    return rows


# ---------------------------
# TENANT USERS (Graph + DB) + SYNC
# ---------------------------
@app.get("/tenant-users", response_model=List[schemas.TenantUserBase])
async def get_tenant_users(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
    authorization: Optional[str] = Header(None),
):
    """
    If Authorization Bearer token provided -> fetch from Graph (live list).
    Else fallback to local DB.
    """
    if authorization and authorization.startswith("Bearer "):
        token = authorization.split(" ")[1]
        try:
            import httpx

            headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
            async with httpx.AsyncClient() as client:
                resp = await client.get(
                    "https://graph.microsoft.com/v1.0/users?$select=displayName,mail,userPrincipalName&$top=50",
                    headers=headers,
                )

            if resp.status_code == 200:
                data = resp.json()
                graph_users = []
                for u in data.get("value", []):
                    email = (u.get("mail") or u.get("userPrincipalName") or "").strip()
                    if not email:
                        continue
                    name = u.get("displayName") or email
                    graph_users.append({"email": email, "name": name})
                return graph_users

            print(f"Graph Users Fetch Failed: {resp.status_code} - {resp.text}")
        except Exception as e:
            print(f"Graph Users Implementation Error: {e}")

    # DB fallback
    users = db.query(models.TenantUser).filter(models.TenantUser.is_active == 1).all()
    return users


@app.post("/tenant-users/sync")
async def sync_tenant_users(
    db: Session = Depends(get_db),
    current_user_email: str = Depends(get_current_user),
    authorization: Optional[str] = Header(None),
):
    """
    Sync tenant users from Graph into local DB (upsert).
    NOTE: This is a tenant-level table. We overwrite rows for simplicity.
    """
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing token")

    token = authorization.split(" ")[1]

    import httpx

    headers = {"Authorization": f"Bearer {token}", "ConsistencyLevel": "eventual"}
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            "https://graph.microsoft.com/v1.0/users?$select=displayName,mail,userPrincipalName&$top=999",
            headers=headers,
        )

    if resp.status_code != 200:
        raise HTTPException(resp.status_code, resp.text)

    data = resp.json().get("value", [])

    # simple "replace all" approach (works fine for small tenants)
    db.query(models.TenantUser).delete()

    inserted = 0
    for u in data:
        email = (u.get("mail") or u.get("userPrincipalName") or "").lower().strip()
        if not email:
            continue
        name = u.get("displayName") or email
        db.add(models.TenantUser(email=email, name=name, tenant_id="default", is_active=1))
        inserted += 1

    db.commit()

    write_audit(
        db,
        owner_email=current_user_email,
        action="SYNC_TENANT_USERS",
        details={"count": inserted},
    )

    return {"ok": True, "count": inserted}



# Microsoft Login APIS

@app.get("/auth/microsoft/login")
def microsoft_login():
    params = {
        "client_id": CLIENT_ID,
        "response_type": "code",
        "redirect_uri": REDIRECT_URI,
        "response_mode": "query",
        "scope": SCOPE,
    }
    url = AUTH_URL + "?" + urlencode(params)
    return RedirectResponse(url)


@app.get("/api/auth/microsoft/callback")
async def microsoft_callback(code: str):
    data = {
        "client_id": CLIENT_ID,
        "scope": SCOPE,
        "code": code,
        "redirect_uri": REDIRECT_URI,
        "grant_type": "authorization_code",
        "client_secret": CLIENT_SECRET,
    }

    async with httpx.AsyncClient() as client:
        token_resp = await client.post(TOKEN_URL, data=data)
        if token_resp.status_code != 200:
            return JSONResponse({"error": "Failed to get token", "details": token_resp.json()})
        
        token_data = token_resp.json()
        access_token = token_data.get("access_token")
        refresh_token = token_data.get("refresh_token")
        expires_in = token_data.get("expires_in")
        token_expires_at = datetime.now(timezone.utc) + timedelta(seconds=expires_in) if expires_in else None
        

        headers = {"Authorization": f"Bearer {access_token}"}
        user_resp = requests.get(USERINFO_URL, headers=headers)
        if user_resp.status_code != 200:
            return JSONResponse({"error": "Failed to get user info", "details": user_resp.json()})
        
        user_data = user_resp.json()
        owner_email_raw = user_data.get("mail") or user_data.get("userPrincipalName")
        if not owner_email_raw:
            raise HTTPException(status_code=400, detail="Email not found in Microsoft profile")

        owner_email = owner_email_raw.strip().lower()
        username = user_data.get("displayName")

    await userdbhandler.save_microsoft_tokens_to_db(owner_email, username, access_token, refresh_token, token_expires_at)

    response = RedirectResponse(url="https://kaylyn-unoppugned-vertie.ngrok-free.dev")

    # 🔐 HTTP-only cookie (NO JS access)
    response.set_cookie(
        key="session_user",
        value=owner_email,
        httponly=True,
        secure=True,   # True in production (HTTPS)
        samesite="none",
        max_age=60 * 60 * 24  # 1 day
    )

    return response

@app.get("/auth/me")
def get_logged_in_user(request: Request):
    email = request.cookies.get("session_user")

    if not email:
        raise HTTPException(status_code=401, detail="Not authenticated")

    return {
        "email": email,
        "loginType": "microsoft"
    }


@app.post("/auth/logout")
def logout():
    response = JSONResponse({"status": "logged out"})
    response.delete_cookie(
        key="session_user",
        httponly=True,
        secure=True,
        samesite="none",
        path="/"
    )

    return response

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app=app, host="127.0.0.1", port=8000, reload=False)

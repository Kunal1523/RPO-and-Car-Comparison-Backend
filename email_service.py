# import pandas as pd
# import io
# import base64
# import httpx
# from typing import List, Dict

# # Settings not needed for Graph Delegate flow as we use passed token
# # from .config import get_settings
# # settings = get_settings()

# async def create_plan_excel(data: dict) -> bytes:
#     """
#     Converts plan data (regulationCells) into an Excel file.
#     Structure: Regulation | Values...
#     """
#     rows = []
#     reg_cells = data.get("regulationCells", {})
    
#     for reg, values in reg_cells.items():
#         # Clean up values roughly
#         row = {"Regulation": reg}
#         for i, val in enumerate(values):
#             row[f"Value_{i+1}"] = val
#         rows.append(row)
        
#     df = pd.DataFrame(rows)
    
#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Final Plan')
        
#     return output.getvalue()

import io
import base64
import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font,PatternFill
from openpyxl.utils import get_column_letter

async def create_plan_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Plan"

    regulation_cells = data["regulationCells"]
    reg_order = data["regOrder"]

    months = [4,5,6,7,8,9,10,11,12,1,2,3]

    quarter_map = {
        "Q1": [4,5,6],
        "Q2": [7,8,9],
        "Q3": [10,11,12],
        "Q4": [1,2,3]
    }

    # 🔥 Extract FYs dynamically
    fys = sorted({key.split("|")[1] for key in regulation_cells.keys()})

    # ================= HEADER =================
    ws.merge_cells("A1:A3")
    ws.cell(row=1, column=1, value="Regulation").alignment = Alignment(
        vertical="center", horizontal="center"
    )

    col = 2
    for fy in fys:
        fy_start_col = col
        fy_months = len(months)

        # FY Header
        ws.merge_cells(
            start_row=1,
            start_column=fy_start_col,
            end_row=1,
            end_column=fy_start_col + fy_months - 1
        )
        ws.cell(row=1, column=fy_start_col, value=fy).alignment = Alignment(horizontal="center")

        # Quarter + Month headers
        for q, m_list in quarter_map.items():
            ws.merge_cells(
                start_row=2,
                start_column=col,
                end_row=2,
                end_column=col + len(m_list) - 1
            )
            ws.cell(row=2, column=col, value=q).alignment = Alignment(horizontal="center")

            for m in m_list:
                ws.cell(row=3, column=col, value=m)
                col += 1

    # ================= DATA =================
    row = 4
    for reg in reg_order:
        ws.cell(row=row, column=1, value=reg)

        col = 2
        for fy in fys:
            for m in months:
                key = f"{reg}|{fy}|{m}"
                values = regulation_cells.get(key, [])
                ws.cell(row=row, column=col, value=", ".join(values) if values else "")
                col += 1

        row += 1

    # ================= EXPORT =================
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()




GRAPH_SEND_MAIL_URL = "https://graph.microsoft.com/v1.0/me/sendMail"


async def send_email_via_graph(
    access_token: str,
    recipients: list[str],
    subject: str,
    body: str,
    attachment_bytes: bytes,
):
    message = {
        "message": {
            "subject": subject,
            "body": {
                "contentType": "HTML",
                "content": body,
            },
            "toRecipients": [
                {"emailAddress": {"address": email}} for email in recipients
            ],
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": "FinalPlan.xlsx",
                    "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "contentBytes": base64.b64encode(attachment_bytes).decode("utf-8"),
                }
            ],
        },
        "saveToSentItems": True,
    }

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
    }

    async with httpx.AsyncClient() as client:
        resp = await client.post(GRAPH_SEND_MAIL_URL, json=message, headers=headers)

    if resp.status_code not in (202, 200):
        raise Exception(f"Graph sendMail failed: {resp.text}")

# async def send_email_via_graph(
#     token: str,
#     recipients: List[str], 
#     subject: str, 
#     body: str, 
#     attachment_bytes: bytes,
#     filename: str = "FinalPlan.xlsx"
# ):
#     """
#     Sends email using Microsoft Graph API (me/sendMail).
#     Requires a valid Delegated Access Token with Mail.Send scope.
#     """
    
#     # 1. Prepare Attachment
#     b64_content = base64.b64encode(attachment_bytes).decode("utf-8")
    
#     email_msg = {
#         "message": {
#             "subject": subject,
#             "body": {
#                 "contentType": "Text",
#                 "content": body
#             },
#             "toRecipients": [
#                 {"emailAddress": {"address": email}} for email in recipients
#             ],
#             "attachments": [
#                 {
#                     "@odata.type": "#microsoft.graph.fileAttachment",
#                     "name": filename,
#                     "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     "contentBytes": b64_content
#                 }
#             ]
#         },
#         "saveToSentItems": "true"
#     }

#     headers = {
#         "Authorization": f"Bearer {token}",
#         "Content-Type": "application/json"
#     }

#     async with httpx.AsyncClient() as client:
#         resp = await client.post(
#             "https://graph.microsoft.com/v1.0/me/sendMail",
#             json=email_msg,
#             headers=headers
#         )
        
#         if resp.status_code != 202:
#             print(f"Graph API Error: {resp.text}")
#             raise Exception(f"Failed to send email via Graph: {resp.status_code} - {resp.text}")

#     print(f"Email sent successfully via Graph API to {recipients}")
